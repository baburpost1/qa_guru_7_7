import os
from openpyxl import load_workbook
from conftest import RESOURCE_PATH


# TODO оформить в тест, добавить ассерты и использовать универсальный путь


def test_xlsx():
    workbook = load_workbook(os.path.join(RESOURCE_PATH, 'file_example_XLSX_50.xlsx'))
    sheet = workbook.active
    print(sheet.cell(row=3, column=2).value)
    sheet_cell = sheet.cell(row=3, column=2).value
    assert sheet_cell == "Mara", "Пересечение столбца и колонки не равно ОЗ"
